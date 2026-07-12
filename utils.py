import os
import shutil
import sqlite3
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font

def generate_report(report_type, year=None, month=None):
    conn = sqlite3.connect('inventory.db')
    if report_type == 'month':
        date_filter = f"strftime('%m', i.issue_date) = '{month:02d}' AND strftime('%Y', i.issue_date) = '{year}'"
        filename = f"report_{year}_{month:02d}.xlsx"
    elif report_type == 'year':
        date_filter = f"strftime('%Y', i.issue_date) = '{year}'"
        filename = f"report_{year}.xlsx"
    else:
        date_filter = "1=1"
        filename = f"stock_report_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    if report_type in ['month', 'year']:
        query = f"""
            SELECT 
                d.name AS 'Отдел',
                e.name AS 'Сотрудник',
                m.model AS 'Модель',
                i.quantity AS 'Количество',
                i.issue_date AS 'Дата',
                i.notes AS 'Примечание'
            FROM issues i
            JOIN departments d ON i.department_id = d.id
            JOIN employees e ON i.employee_id = e.id
            JOIN models m ON i.model_id = m.id
            WHERE {date_filter}
            ORDER BY d.name, i.issue_date DESC
        """
        df = pd.read_sql_query(query, conn)
    else:
        query = """
            SELECT 
                m.model AS 'Модель',
                COALESCE(SUM(inc.quantity), 0) AS 'Приход',
                COALESCE(SUM(iss.quantity), 0) AS 'Выдано',
                COALESCE(SUM(inc.quantity), 0) - COALESCE(SUM(iss.quantity), 0) AS 'Остаток'
            FROM models m
            LEFT JOIN incoming inc ON m.id = inc.model_id
            LEFT JOIN issues iss ON m.id = iss.model_id AND iss.status = 'issued'
            GROUP BY m.id
            ORDER BY m.model
        """
        df = pd.read_sql_query(query, conn)
    conn.close()
    if df.empty:
        return None
    df.to_excel(filename, index=False, engine='openpyxl')
    return filename

def generate_custom_report(month, year):
    os.makedirs('templates', exist_ok=True)
    template_path = os.path.join('templates', 'spisanie_template.xlsx')
    
    if not os.path.exists(template_path):
        create_example_template()
        return None
    
    conn = sqlite3.connect('inventory.db')
    date_filter = f"strftime('%m', i.issue_date) = '{month:02d}' AND strftime('%Y', i.issue_date) = '{year}'"
    query = f"""
        SELECT 
            d.name AS 'Отдел',
            m.model AS 'Картридж',
            e.name AS 'Фамилия'
        FROM issues i
        JOIN departments d ON i.department_id = d.id
        JOIN employees e ON i.employee_id = e.id
        JOIN models m ON i.model_id = m.id
        WHERE {date_filter}
        ORDER BY d.name
    """
    df = pd.read_sql_query(query, conn)
    conn.close()
    if df.empty:
        return None
    
    month_names = ['', 'Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
                   'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь']
    month_name = month_names[month] if 1 <= month <= 12 else str(month)
    filename = f"списание_{month_name}_{year}.xlsx"
    shutil.copy(template_path, filename)
    wb = load_workbook(filename)
    ws = wb.active
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and 'Период списания:' in str(cell.value):
                cell.value = f"Период списания: {month_name} {year}"
    
    headers = ['Отдел', 'Картридж', 'Фамилия']
    start_row, start_col = None, None
    for row_idx, row in enumerate(ws.iter_rows(), 1):
        for col_idx, cell in enumerate(row, 1):
            if cell.value in headers:
                start_row, start_col = row_idx + 1, col_idx
                break
        if start_row:
            break
    if start_row is None:
        start_row, start_col = 4, 1
    
    for row in ws.iter_rows(min_row=start_row, max_col=start_col+2):
        for cell in row:
            if cell.row > 2:
                cell.value = None
    
    current_row = start_row
    for _, row in df.iterrows():
        ws.cell(row=current_row, column=start_col, value=row['Отдел'])
        ws.cell(row=current_row, column=start_col+1, value=row['Картридж'])
        ws.cell(row=current_row, column=start_col+2, value=row['Фамилия'])
        current_row += 1
    
    ws.cell(row=current_row, column=start_col, value="Итого:")
    ws.cell(row=current_row, column=start_col+1, value=f"{len(df)} шт.")
    wb.save(filename)
    return filename

def create_example_template():
    os.makedirs('templates', exist_ok=True)
    template_path = os.path.join('templates', 'spisanie_template.xlsx')
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Лист1"
    ws['A1'] = "Период списания:"
    ws['C1'] = "Июнь 2026"
    ws['A3'] = "Отдел"
    ws['B3'] = "Картридж"
    ws['C3'] = "Фамилия"
    ws['D3'] = "Подпись"
    for cell in ws['A3:D3'][0]:
        cell.font = Font(bold=True)
    ws['A6'] = "Ответственный 1"
    ws['C6'] = "Ответственный 2"
    wb.save(template_path)
    return template_path
    
   

def backup_database():
    """Экспортирует все данные в Excel-файл"""
    conn = sqlite3.connect('inventory.db')
    
    # Список таблиц для экспорта
    tables = ['departments', 'employees', 'models', 'incoming', 'issues']
    
    # Создаем папку для бекапов, если её нет
    os.makedirs('backups', exist_ok=True)
    
    # Имя файла с датой
    filename = f"backups/backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        for table in tables:
            try:
                df = pd.read_sql_query(f"SELECT * FROM {table}", conn)
                df.to_excel(writer, sheet_name=table, index=False)
            except Exception as e:
                print(f"Ошибка при экспорте таблицы {table}: {e}")
    
    conn.close()
    return filename

def restore_from_backup(filepath):
    """Восстанавливает базу данных из Excel-файла"""
    if not os.path.exists(filepath):
        return False, "Файл не найден"
    
    # Читаем все листы из Excel
    try:
        xls = pd.ExcelFile(filepath)
        sheet_names = xls.sheet_names
        
        conn = sqlite3.connect('inventory.db')
        cur = conn.cursor()
        
        # Отключаем проверку внешних ключей на время восстановления
        cur.execute("PRAGMA foreign_keys = OFF")
        
        # Список таблиц и их колонок (для правильного порядка)
        table_columns = {
            'departments': ['id', 'name'],
            'employees': ['id', 'name', 'department_id'],
            'models': ['id', 'model'],
            'incoming': ['id', 'model_id', 'quantity', 'receive_date', 'notes'],
            'issues': ['id', 'model_id', 'department_id', 'employee_id', 
                       'quantity', 'issue_date', 'status', 'notes']
        }
        
        # Очищаем существующие таблицы
        for table in table_columns.keys():
            if table in sheet_names:
                cur.execute(f"DELETE FROM {table}")
        
        # Восстанавливаем данные из каждого листа
        for table, columns in table_columns.items():
            if table not in sheet_names:
                continue
            
            df = pd.read_excel(filepath, sheet_name=table)
            
            # Убеждаемся, что все колонки есть
            for col in columns:
                if col not in df.columns:
                    df[col] = None
            
            # Приводим порядок колонок к нужному
            df = df[columns]
            
            # Заменяем NaN на None для SQLite
            df = df.where(pd.notna(df), None)
            
            # Вставляем данные
            placeholders = ','.join(['?' for _ in columns])
            col_names = ','.join(columns)
            
            for _, row in df.iterrows():
                try:
                    cur.execute(f"INSERT INTO {table} ({col_names}) VALUES ({placeholders})", 
                                tuple(row))
                except Exception as e:
                    print(f"Ошибка при вставке в {table}: {e}")
        
        # Включаем проверку внешних ключей обратно
        cur.execute("PRAGMA foreign_keys = ON")
        conn.commit()
        conn.close()
        
        return True, "База успешно восстановлена"
        
    except Exception as e:
        return False, f"Ошибка при восстановлении: {str(e)}"
    